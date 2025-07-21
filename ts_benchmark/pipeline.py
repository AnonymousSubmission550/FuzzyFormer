# -*- coding: utf-8 -*-
from dataclasses import dataclass
from functools import reduce
from operator import and_
from typing import List, Dict, Type, Optional
import pandas as pd
from ts_benchmark.data.data_source import (
    LocalForecastingDataSource,
    DataSource,
)
from ts_benchmark.data.suites.global_storage import GlobalStorageDataServer
from ts_benchmark.evaluation.evaluate_model import eval_model
from ts_benchmark.models import get_models
from ts_benchmark.recording import save_log
from ts_benchmark.utils.parallel import ParallelBackend
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

DEFAULT_TRANSFORMER_BASED_HYPER_PARAMS = {
    "top_k": 5,
    "enc_in": 1,
    "dec_in": 1,
    "c_out": 1,
    "e_layers": 2,
    "d_layers": 1,
    "d_model": 512,
    "d_ff": 2048,
    "embed": "timeF",
    "freq": "h",
    "lradj": "type1",
    "moving_avg": 25,
    "num_kernels": 6,
    "factor": 1,
    "n_heads": 8,
    "seg_len": 6,
    "win_size": 2,
    "activation": "gelu",
    "output_attention": 0,
    "patch_len": 16,
    "stride": 8,
    "dropout": 0.1,
    "batch_size": 32,
    "lr": 0.0001,
    "num_epochs": 10,
    "num_workers": 0,
    "loss": "MSE",
    "itr": 1,
    "distil": True,
    "patience": 3,
    "p_hidden_dims": [128, 128],
    "p_hidden_layers": 2,
    "mem_dim": 32,
    "conv_kernel": [12, 16],
    "anomaly_ratio": 1.0,
    "down_sampling_windows": 2,
    "channel_independence": True,
    "down_sampling_layers": 3,
    "down_sampling_method": "avg",
    "decomp_method": "moving_avg",
    "use_norm": True,
}


@dataclass
class DatasetInfo:
    # the possible values of the meta-info field 'size'
    size_value: List
    # the class of data source for this dataset
    datasrc_class: Type[DataSource]


PREDEFINED_DATASETS = {
    "large_forecast": DatasetInfo(size_value=["large", "small"], datasrc_class=LocalForecastingDataSource),
    "small_forecast": DatasetInfo(size_value=["small"], datasrc_class=LocalForecastingDataSource),
    "user_forecast": DatasetInfo(size_value=["user"], datasrc_class=LocalForecastingDataSource),
}


def filter_data(
    metadata: pd.DataFrame,
    size_value: List[str],
    feature_dict: Optional[Dict] = None,
) -> List[str]:
    """
    Filters the dataset based on given filters

    :param metadata: The meta information DataFrame.
    :param size_value: The allowed values of the 'size' meta-info field.
    :param feature_dict: A dictionary of filters where each key is a meta-info field
        and the corresponding value is the field value to keep. If None is given,
        no extra filter is applied.
    :return:
    """
    # Remove items with a value of None in feature_dict
    feature_dict = {k: v for k, v in feature_dict.items() if v is not None}

    # Use the reduce and and_ functions to filter data file names that meet the criteria
    filt_metadata = metadata
    if feature_dict is not None:
        filt_metadata = metadata[reduce(and_, (metadata[k] == v for k, v in feature_dict.items()))]
    filt_metadata = filt_metadata[filt_metadata["size"].isin(size_value)]

    return filt_metadata["file_name"].tolist()


def _get_model_names(model_names: List[str]):
    """
    Rename models if there exists duplications.

    If a model A appears multiple times in the list, each appearance will be renamed to
    `A`, `A_1`, `A_2`, ...

    :param model_names: A list of model names.
    :return: The renamed list of model names.
    """
    s = pd.Series(model_names)
    cumulative_counts = s.groupby(s).cumcount()
    return [f"{model_name}_{cnt}" if cnt > 0 else model_name for model_name, cnt in zip(model_names, cumulative_counts)]


def pipeline(data_config: dict, model_config: dict, evaluation_config: dict) -> List[str]:
    """
    Execute the benchmark pipeline process

    The pipline includes loading data, building models, evaluating models, and generating reports.

    :param data_config: Configuration for data loading.
    :param model_config: Configuration for model construction.
    :param evaluation_config: Configuration for model evaluation.
    """
    # prepare data
    # TODO: move these code into the data module, after the pipeline interface is unified
    dataset_name_list = data_config.get("data_set_name", ["small_forecast"])
    if not dataset_name_list:
        dataset_name_list = ["small_forecast"]
    if isinstance(dataset_name_list, str):
        dataset_name_list = [dataset_name_list]
    for dataset_name in dataset_name_list:
        if dataset_name not in PREDEFINED_DATASETS:
            raise ValueError(f"Unknown dataset {dataset_name}.")

    data_src_type = PREDEFINED_DATASETS[dataset_name_list[0]].datasrc_class
    if not all(PREDEFINED_DATASETS[dataset_name].datasrc_class is data_src_type for dataset_name in dataset_name_list):
        raise ValueError("Not supporting different types of data sources.")

    data_src: DataSource = PREDEFINED_DATASETS[dataset_name_list[0]].datasrc_class()
    data_name_list = data_config.get("data_name_list", None)
    if not data_name_list:
        data_name_list = []
        for dataset_name in dataset_name_list:
            size_value = PREDEFINED_DATASETS[dataset_name].size_value
            feature_dict = data_config.get("feature_dict", None)
            data_name_list.extend(filter_data(data_src.dataset.metadata, size_value, feature_dict=feature_dict))

    data_name_list = list(set(data_name_list))
    if not data_name_list:
        raise ValueError("No dataset specified.")
    data_src.load_series_list(data_name_list)
    data_server = GlobalStorageDataServer(data_src, ParallelBackend())
    data_server.start_async()

    # modeling
    model_factory_list = get_models(model_config)
    model_name = model_factory_list[0].model_name
    result_list = [eval_model(model_factory, data_name_list, evaluation_config) for model_factory in model_factory_list]

    # save
    result_itr = result_list[0].collect()
    result_df = list(result_itr)
    data_set = data_name_list[0]
    lr = (
        model_config.get("models", [{}])[0]
        .get("model_hyper_params", {})
        .get("lr", DEFAULT_TRANSFORMER_BASED_HYPER_PARAMS["lr"])
    )
    epo_num = (
        model_config.get("models", [{}])[0]
        .get("model_hyper_params", {})
        .get("num_epochs", DEFAULT_TRANSFORMER_BASED_HYPER_PARAMS["num_epochs"])
    )
    Batch_size = (
        model_config.get("models", [{}])[0]
        .get("model_hyper_params", {})
        .get("batch_size", DEFAULT_TRANSFORMER_BASED_HYPER_PARAMS["batch_size"])
    )
    seq_len = model_config.get("models", [{}])[0].get("model_hyper_params", {}).get("seq_len", 104)
    horizon = model_config.get("models", [{}])[0].get("model_hyper_params", {}).get("horizon", 24)
    seed = evaluation_config.get("strategy_args", {}).get("seed", 2021)
    MSE = round(result_df[0]["mse_norm"].values[0], 5)
    MAE = round(result_df[0]["mae_norm"].values[0], 5)
    train_time = round(result_df[0]["fit_time"].values[0], 6)
    test_time = round(result_df[0]["inference_time"].values[0], 6)
    varpi_1 = model_config.get("models", [{}])[0].get("model_hyper_params", {}).get("varpi_1", 0.1)
    varpi_2 = model_config.get("models", [{}])[0].get("model_hyper_params", {}).get("varpi_2", 1)
    dropout = model_config.get("models", [{}])[0].get("model_hyper_params", {}).get("dropout", 1)
    # 读取参数数量
    with open('params.txt', 'r') as f:
        line = f.readline()
        total_params = int(line.strip().split(':')[-1].replace(',', ''))
    save_data = [
        model_name,
        data_set,
        str(lr),
        str(varpi_1),
        str(varpi_2),
        str(dropout),
        str(epo_num),
        str(Batch_size),
        str(seq_len),
        str(horizon),
        str(seed),
        str(MSE),
        str(MAE),
        str(train_time),
        str(test_time),
        str(total_params),
    ]
    print("**********************************************")
    print(f"Mse is {MSE}")
    print(f"Mae is {MAE}")
    print("**********************************************")
    workbook = load_workbook("result.xlsx")
    worksheet = workbook.active
    worksheet.append(save_data)
    font_style = Font(name="Times New Roman", size=14, color="000000")
    alignment = Alignment(horizontal="center", vertical="center")
    for cell in worksheet[worksheet.max_row]:
        cell.font = font_style
        cell.alignment = alignment
    workbook.save("result.xlsx")
    return MSE, MAE
